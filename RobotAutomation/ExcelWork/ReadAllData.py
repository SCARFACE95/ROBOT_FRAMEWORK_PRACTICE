#Import package

import openpyxl

# Load Workbook

vk = openpyxl.load_workbook("/Users/Robot_framework_projects/RobotAutomation/TesData.xlsx")
sh = vk['Sheet1']

#Find out how many rows we are having data on sheet
rows = sh.max_row

#Find out how many collums we are having data on sheet
columns = sh.max_column

print(f"Total Rows are -> {rows}")
print(f"Total Columns are -> {columns}")


# for i in range(1, rows+1):
#     for j in range(1, columns+1):
#         c=sh.cell(i,j)
#         print(c.value)


#Another approach

for r in sh['A1':'C4']:
    for c in r:
        print(c.value)